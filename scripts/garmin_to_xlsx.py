#!/usr/bin/env python3
"""
Sync Garmin data into workspace xlsx file.

Primary flow:
1) Optionally refresh ~/.clawdbot/.garmin-cache.json via garmin-sync.py
2) Read cache JSON and extract daily + workout data
3) Update row by date in 训练饮食记录表.xlsx using header names
"""

from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
from collections import Counter, OrderedDict
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.formula.translate import Translator


DEFAULT_HEADERS = [
    "日期",
    "星期",
    "周次",
    "体重(kg)",
    "体脂率(%)",
    "目标体重(kg)",
    "与目标差值(kg)",
    "7日均重(kg)",
    "目标体脂率(%)",
    "当日完成度(%)",
    "静息心率(次/分)",
    "步数",
    "总消耗卡路里(大卡)",
    "睡眠时长(小时)",
    "是否训练",
    "训练部位",
    "训练时长(分钟)",
    "训练动作",
]

BEIJING_TZ = timezone(timedelta(hours=8))
SYNC_FIELDS = [
    "体重(kg)",
    "静息心率(次/分)",
    "步数",
    "总消耗卡路里(大卡)",
    "睡眠时长(小时)",
    "训练时长(分钟)",
    "训练动作",
    "训练部位",
    "是否训练",
]
WORKOUT_TYPE_MAP = {
    "running": "跑步",
    "walking": "步行",
    "cycling": "骑行",
    "indoor_cardio": "有氧",
    "strength_training": "力量",
    "swimming": "游泳",
    "hiking": "徒步",
}
MUSCLE_GROUP_CYCLE = ["胸", "背", "肩膀", "臀腿"]
# 用户当前计划默认锚点：2026-03-18 作为一个循环起点（胸）
CYCLE_ANCHOR_DATE = "2026-03-18"
CATEGORY_RULES = {
    "胸": ("BENCH", "CHEST", "PUSH_UP", "PEC", "FLY", "DIP"),
    "背": ("PULL_UP", "ROW", "LAT", "PULLDOWN", "DEADLIFT", "BACK_EXTENSION"),
    "肩膀": ("SHOULDER", "OVERHEAD_PRESS", "LATERAL_RAISE", "FRONT_RAISE", "REAR_DELT", "UPRIGHT_ROW"),
    "臀腿": ("SQUAT", "LUNGE", "LEG_PRESS", "LEG_EXTENSION", "LEG_CURL", "CALF", "GLUTE", "HIP_THRUST", "STEP_UP"),
}
NAME_KEYWORD_RULES = {
    "胸": ("卧推", "上斜", "下斜", "飞鸟", "夹胸", "俯卧撑", "chest", "bench", "pec"),
    "背": ("划船", "引体", "下拉", "硬拉", "背阔", "row", "pull", "lat", "deadlift"),
    "肩膀": ("肩推", "推举", "侧平举", "前平举", "后束", "面拉", "shoulder", "press", "deltoid"),
    "臀腿": ("深蹲", "腿举", "腿弯举", "腿伸展", "弓步", "臀桥", "臀推", "小腿", "squat", "leg", "glute", "hip"),
}
EMPTY_TEXT_TOKENS = {"", "/", "／", "-", "--", "—", "无", "none", "n/a", "na", "null"}
MEAL_HEADERS = ("早餐", "午餐", "晚餐", "加餐")
SUMMARY_HEADERS = ("热量和营养成分分析", "饮食总结")
CALORIE_HEADERS = ("总热量(大卡)", "总热量摄入(大卡)")
PROTEIN_HEADERS = ("蛋白质摄入(g)",)
CARB_HEADERS = ("碳水摄入(g)",)
FAT_HEADERS = ("脂肪摄入(g)",)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Sync Garmin cache data to xlsx.")
    parser.add_argument(
        "--xlsx",
        default=str(default_xlsx_path()),
        help="Target xlsx path. Default: workspace/训练饮食记录表.xlsx",
    )
    parser.add_argument(
        "--cache",
        default=str(Path.home() / ".clawdbot" / ".garmin-cache.json"),
        help="Garmin cache JSON path.",
    )
    parser.add_argument(
        "--date",
        default=None,
        help="Target date in YYYY-MM-DD. Default: cache date, then today (Beijing).",
    )
    parser.add_argument(
        "--no-sync",
        action="store_true",
        help="Do not call garmin-sync.py before writing xlsx.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print planned updates without writing file.",
    )
    parser.add_argument(
        "--clear-missing",
        action="store_true",
        help="Clear sync fields when value is missing for the target date.",
    )
    parser.add_argument(
        "--write-summary-to-remark",
        action="store_true",
        help="Also write workout summary to '备注' column if present.",
    )
    parser.add_argument(
        "--summary-max-len",
        type=int,
        default=120,
        help="Max characters for workout summary text.",
    )
    parser.add_argument(
        "--query-muscle-group",
        action="store_true",
        help="Query inferred muscle group for target date and print JSON without writing xlsx.",
    )
    parser.add_argument(
        "--no-write-inference-reason",
        action="store_true",
        help="Do not write muscle-group inference reason into '备注'.",
    )
    return parser.parse_args()


def default_xlsx_path() -> Path:
    script = Path(__file__).resolve()
    workspace_guess = script.parents[3] / "训练饮食记录表.xlsx"
    if workspace_guess.exists():
        return workspace_guess
    return Path.home() / ".openclaw" / "workspace" / "训练饮食记录表.xlsx"


def run_sync_if_needed(skip_sync: bool, target_date: Optional[str] = None) -> None:
    if skip_sync:
        return

    sync_script = Path(__file__).resolve().parent / "garmin-sync.py"
    if not sync_script.exists():
        raise FileNotFoundError(f"Missing sync script: {sync_script}")

    cmd = [sys.executable, str(sync_script)]
    if target_date:
        # Keep cache date aligned with requested write date.
        cmd.extend(["--date", target_date])

    result = subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        check=False,
    )
    if result.returncode != 0:
        stderr = (result.stderr or "").strip()
        stdout = (result.stdout or "").strip()
        msg = stderr or stdout or "unknown error"
        raise RuntimeError(f"garmin-sync.py failed: {msg}")


def load_cache(cache_path: Path) -> Dict:
    if not cache_path.exists():
        raise FileNotFoundError(f"Cache not found: {cache_path}")
    with cache_path.open("r", encoding="utf-8") as f:
        return json.load(f)


def pick_target_date(cache: Dict, cli_date: Optional[str]) -> str:
    if cli_date:
        datetime.strptime(cli_date, "%Y-%m-%d")
        return cli_date
    cache_date = cache.get("date")
    if cache_date:
        datetime.strptime(cache_date, "%Y-%m-%d")
        return cache_date
    return datetime.now(BEIJING_TZ).strftime("%Y-%m-%d")


def find_header_row(ws) -> int:
    for row in range(1, min(ws.max_row, 20) + 1):
        for col in range(1, min(ws.max_column, 40) + 1):
            if ws.cell(row, col).value == "日期":
                return row
    return 3


def ensure_workbook(xlsx_path: Path):
    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        return wb, wb[wb.sheetnames[0]]

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(1, 1).value = "Garmin 自动同步生成"
    ws.cell(2, 1).value = "可按需补充其它列和公式"
    for idx, header in enumerate(DEFAULT_HEADERS, start=1):
        ws.cell(3, idx).value = header
    return wb, ws


def get_header_map(ws, header_row: int) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(header_row, col).value
        if isinstance(header, str) and header.strip():
            header_map[header.strip()] = col
    return header_map


def first_existing_col(header_map: Dict[str, int], aliases: Iterable[str]) -> Optional[int]:
    for name in aliases:
        col = header_map.get(name)
        if col:
            return col
    return None


def normalize_text(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.lower() in EMPTY_TEXT_TOKENS:
        return None
    return text


def parse_metric_from_summary(summary_text: str, keywords: Iterable[str]) -> Optional[float]:
    if not summary_text:
        return None
    escaped = [re.escape(k) for k in keywords]
    if not escaped:
        return None
    segment_pattern = rf"(?:{'|'.join(escaped)})[^\n]*"
    for match in re.finditer(segment_pattern, summary_text, flags=re.IGNORECASE):
        segment = match.group(0)
        median = re.search(r"中位(?:值)?(?:约)?\s*([0-9]+(?:\.[0-9]+)?)", segment)
        if median:
            return float(median.group(1))
        ranged = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*[-–~～至到]\s*([0-9]+(?:\.[0-9]+)?)", segment)
        if ranged:
            low = float(ranged.group(1))
            high = float(ranged.group(2))
            return (low + high) / 2.0
        number = re.search(r"([0-9]+(?:\.[0-9]+)?)", segment)
        if number:
            return float(number.group(1))
    return None


def parse_nutrition_from_summary(summary_text: Optional[str]) -> Dict[str, float]:
    text = normalize_text(summary_text)
    if not text:
        return {}
    return {
        "calories": parse_metric_from_summary(text, ("总热量", "热量", "kcal", "大卡", "千卡")),
        "protein": parse_metric_from_summary(text, ("蛋白质", "protein")),
        "carb": parse_metric_from_summary(text, ("碳水", "碳水化合物", "carb")),
        "fat": parse_metric_from_summary(text, ("脂肪", "fat")),
    }


def maybe_write_cell(ws, row: int, col: Optional[int], value: object) -> bool:
    if not col or value is None:
        return False
    cell = ws.cell(row, col)
    existing = cell.value
    if normalize_text(existing) is not None:
        return False
    cell.value = value
    return True


def read_number(value: object) -> Optional[float]:
    if isinstance(value, (int, float)):
        return float(value)
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.startswith("="):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def build_nutrition_summary_text(calories: Optional[float], protein: Optional[float], carb: Optional[float], fat: Optional[float]) -> Optional[str]:
    parts: List[str] = []
    if calories and calories > 0:
        parts.append(f"总热量约{int(round(calories))}kcal")
    if protein and protein > 0:
        parts.append(f"蛋白质约{int(round(protein))}g")
    if carb and carb > 0:
        parts.append(f"碳水约{int(round(carb))}g")
    if fat and fat > 0:
        parts.append(f"脂肪约{int(round(fat))}g")
    if not parts:
        return None
    return "；".join(parts)


def sync_nutrition_from_row(ws, row: int, header_map: Dict[str, int]) -> Tuple[Dict[str, object], Dict[str, object]]:
    written: Dict[str, object] = {}
    meal_values: Dict[str, Optional[str]] = {}
    for header in MEAL_HEADERS:
        col = header_map.get(header)
        meal_values[header] = normalize_text(ws.cell(row, col).value) if col else None
    filled_meals = sum(1 for v in meal_values.values() if v is not None)

    # Support both historical and current column names.
    summary_text = None
    for header in SUMMARY_HEADERS:
        col = header_map.get(header)
        if not col:
            continue
        text = normalize_text(ws.cell(row, col).value)
        if text:
            summary_text = text
            break

    parsed = parse_nutrition_from_summary(summary_text)
    col_to_header = {col: header for header, col in header_map.items()}

    calorie_col = first_existing_col(header_map, CALORIE_HEADERS)
    protein_col = first_existing_col(header_map, PROTEIN_HEADERS)
    carb_col = first_existing_col(header_map, CARB_HEADERS)
    fat_col = first_existing_col(header_map, FAT_HEADERS)

    if maybe_write_cell(ws, row, calorie_col, int(round(parsed["calories"])) if parsed.get("calories") else None):
        written[col_to_header[calorie_col]] = int(round(parsed["calories"]))
    if maybe_write_cell(ws, row, protein_col, int(round(parsed["protein"])) if parsed.get("protein") else None):
        written[col_to_header[protein_col]] = int(round(parsed["protein"]))
    if maybe_write_cell(ws, row, carb_col, int(round(parsed["carb"])) if parsed.get("carb") else None):
        written[col_to_header[carb_col]] = int(round(parsed["carb"]))
    if maybe_write_cell(ws, row, fat_col, int(round(parsed["fat"])) if parsed.get("fat") else None):
        written[col_to_header[fat_col]] = int(round(parsed["fat"]))

    calories_value = read_number(ws.cell(row, calorie_col).value) if calorie_col else None
    protein_value = read_number(ws.cell(row, protein_col).value) if protein_col else None
    carb_value = read_number(ws.cell(row, carb_col).value) if carb_col else None
    fat_value = read_number(ws.cell(row, fat_col).value) if fat_col else None
    summary_target_col = first_existing_col(header_map, SUMMARY_HEADERS)
    if summary_target_col:
        existing_summary = normalize_text(ws.cell(row, summary_target_col).value)
        if not existing_summary:
            generated = summary_text or build_nutrition_summary_text(
                calories=calories_value,
                protein=protein_value,
                carb=carb_value,
                fat=fat_value,
            )
            if generated:
                ws.cell(row, summary_target_col).value = generated
                written[col_to_header[summary_target_col]] = generated

    status = {
        "filled_meals": filled_meals,
        "meal_values": meal_values,
        "summary_present": bool(normalize_text(ws.cell(row, summary_target_col).value)) if summary_target_col else False,
    }
    return written, status


def find_or_create_date_row(ws, header_row: int, header_map: Dict[str, int], date_str: str) -> int:
    date_col = header_map.get("日期", 1)
    for row in range(header_row + 1, ws.max_row + 1):
        if str(ws.cell(row, date_col).value) == date_str:
            return row

    new_row = ws.max_row + 1
    ws.cell(new_row, date_col).value = date_str
    source_row = max(header_row + 1, new_row - 1)
    if source_row < new_row:
        copy_formulas(ws, source_row, new_row)
    return new_row


def copy_formulas(ws, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(source_row, col)
        value = source_cell.value
        if isinstance(value, str) and value.startswith("="):
            target_cell = ws.cell(target_row, col)
            try:
                target_cell.value = Translator(value, origin=source_cell.coordinate).translate_formula(
                    target_cell.coordinate
                )
            except Exception:
                target_cell.value = value


def normalize_weight_kg(raw: Optional[float]) -> Optional[float]:
    if raw is None:
        return None
    try:
        weight = float(raw)
    except (TypeError, ValueError):
        return None

    if weight <= 0:
        return None
    if weight > 300:
        weight = weight / 1000.0
    if weight < 25 or weight > 250:
        return None
    return round(weight, 2)


def workout_date(workout: Dict) -> Optional[str]:
    date_text = workout.get("date")
    if isinstance(date_text, str) and len(date_text) == 10:
        return date_text

    timestamp = workout.get("timestamp")
    if isinstance(timestamp, (int, float)) and timestamp > 0:
        return datetime.fromtimestamp(timestamp, BEIJING_TZ).strftime("%Y-%m-%d")
    return None


def infer_training_area(
    day_workouts: List[Dict],
    workout_names: Iterable[str],
    target_date: str,
) -> Tuple[str, str, str, Dict[str, int]]:
    category_counter: Counter = Counter()
    unknown_count = 0

    for workout in day_workouts:
        counts = workout.get("exercise_category_counts")
        if isinstance(counts, dict):
            for cat, c in counts.items():
                try:
                    n = int(c)
                except (TypeError, ValueError):
                    continue
                if n > 0 and cat:
                    category_counter[str(cat).upper()] += n
        unknown_count += int(workout.get("exercise_unknown_count") or 0)

    if category_counter:
        score = {k: 0 for k in CATEGORY_RULES.keys()}
        for category, count in category_counter.items():
            for group, tokens in CATEGORY_RULES.items():
                if any(token in category for token in tokens):
                    score[group] += count

        best = max(score, key=score.get)
        if score[best] > 0:
            top_codes = ", ".join([f"{k}x{v}" for k, v in category_counter.most_common(4)])
            reason = f"命中Garmin分类码[{top_codes}]"
            if unknown_count > 0:
                reason += f"，另有{unknown_count}组未识别分类"
            return best, reason, "garmin_category", dict(category_counter)

    names = " ".join(workout_names).lower()
    if names:
        score = {k: 0 for k in NAME_KEYWORD_RULES.keys()}
        for group, keywords in NAME_KEYWORD_RULES.items():
            for kw in keywords:
                if kw in names:
                    score[group] += 1
        best = max(score, key=score.get)
        if score[best] > 0:
            return best, f"分类码不足，按动作名关键词推断: {names[:80]}", "name_keyword", {}

    # 名称模糊（如“力量训练”）时按固定循环推断
    anchor = datetime.strptime(CYCLE_ANCHOR_DATE, "%Y-%m-%d").date()
    day = datetime.strptime(target_date, "%Y-%m-%d").date()
    idx = (day - anchor).days % len(MUSCLE_GROUP_CYCLE)
    group = MUSCLE_GROUP_CYCLE[idx]
    return group, f"分类码/动作名不足，按循环推断(锚点{CYCLE_ANCHOR_DATE}={MUSCLE_GROUP_CYCLE[0]})", "cycle_fallback", {}


def workout_type_label(workout: Dict) -> str:
    wtype = workout.get("type")
    if isinstance(wtype, dict):
        key = wtype.get("typeKey")
        if isinstance(key, str):
            return WORKOUT_TYPE_MAP.get(key, key)
    if isinstance(wtype, str):
        return WORKOUT_TYPE_MAP.get(wtype, wtype)
    return ""


def build_workout_summary(day_workouts: List[Dict], max_len: int = 120) -> Optional[str]:
    if not day_workouts:
        return None

    normalized = sorted(
        day_workouts,
        key=lambda w: int(w.get("timestamp") or 0),
    )
    items: List[str] = []
    total_cal = 0
    total_minutes = 0

    for workout in normalized:
        name = str(workout.get("name") or "").strip()
        if not name:
            name = workout_type_label(workout) or "训练"
        minutes = int(round(float(workout.get("duration_minutes") or 0)))
        calories = int(round(float(workout.get("calories") or 0)))
        total_minutes += max(minutes, 0)
        total_cal += max(calories, 0)
        if minutes > 0 and calories > 0:
            items.append(f"{name}{minutes}分/{calories}kcal")
        elif minutes > 0:
            items.append(f"{name}{minutes}分")
        else:
            items.append(name)

    base = "；".join(items[:6])
    suffix = f"（共{len(day_workouts)}项 {total_minutes}分 {total_cal}kcal）"
    text = f"{base}{suffix}" if base else suffix
    if len(text) > max_len:
        text = f"{text[:max_len-1]}…"
    return text


def build_updates(
    cache: Dict,
    target_date: str,
    write_summary_to_remark: bool = False,
    summary_max_len: int = 120,
    write_inference_reason: bool = True,
) -> Dict[str, object]:
    cache_date = cache.get("date")
    cache_match_target = str(cache_date) == str(target_date)

    # Guardrail: never copy yesterday summary into today row when cache date mismatches.
    summary = (cache.get("summary", {}) or {}) if cache_match_target else {}
    sleep = (cache.get("sleep", {}) or {}) if cache_match_target else {}
    weight_info = (cache.get("weight", {}) or {}) if cache_match_target else {}
    workouts = cache.get("workouts", []) or []

    day_workouts = [w for w in workouts if workout_date(w) == target_date]
    names = []
    for w in day_workouts:
        name = w.get("name") or ""
        if isinstance(name, str) and name.strip():
            names.append(name.strip())
    names = list(OrderedDict.fromkeys(names))

    total_minutes = 0
    for w in day_workouts:
        try:
            total_minutes += float(w.get("duration_minutes") or 0)
        except (TypeError, ValueError):
            continue

    weight_kg = normalize_weight_kg(weight_info.get("weight_kg"))
    if weight_kg is None:
        weight_kg = normalize_weight_kg(summary.get("weight_kg"))

    hr = summary.get("heart_rate_resting")
    steps = summary.get("steps")
    calories = summary.get("calories")
    sleep_hours = sleep.get("duration_hours")

    workout_summary = build_workout_summary(day_workouts, max_len=summary_max_len)

    muscle_group = None
    inference_reason = None
    inference_method = None
    category_counts: Dict[str, int] = {}
    if day_workouts:
        muscle_group, inference_reason, inference_method, category_counts = infer_training_area(
            day_workouts=day_workouts,
            workout_names=names,
            target_date=target_date,
        )

    updates: Dict[str, object] = {
        "体重(kg)": weight_kg,
        "静息心率(次/分)": hr if isinstance(hr, (int, float)) and hr > 0 else None,
        "步数": int(steps) if isinstance(steps, (int, float)) and steps > 0 else None,
        "总消耗卡路里(大卡)": int(round(calories)) if isinstance(calories, (int, float)) and calories > 0 else None,
        "睡眠时长(小时)": round(float(sleep_hours), 1) if isinstance(sleep_hours, (int, float)) and sleep_hours > 0 else None,
        "训练时长(分钟)": int(round(total_minutes)) if total_minutes > 0 else None,
        "训练动作": workout_summary or ("、".join(names) if names else None),
        "训练部位": muscle_group,
        "是否训练": "是" if names else None,
    }
    if write_inference_reason and muscle_group:
        reason_parts = []
        if write_summary_to_remark and workout_summary:
            reason_parts.append(f"训练摘要：{workout_summary}")
        reason_parts.append(f"肌群推断：{muscle_group}")
        if inference_reason:
            reason_parts.append(f"依据：{inference_reason}")
        updates["备注"] = "；".join(reason_parts)
    elif write_summary_to_remark:
        updates["备注"] = workout_summary

    updates["_inference_reason"] = inference_reason
    updates["_inference_method"] = inference_method
    updates["_category_counts"] = category_counts
    updates["_cache_date"] = cache_date
    updates["_cache_match_target"] = cache_match_target
    return updates


def apply_updates(
    ws,
    row: int,
    header_map: Dict[str, int],
    updates: Dict[str, object],
    clear_missing: bool = False,
) -> Dict[str, object]:
    written: Dict[str, object] = {}
    ordered_fields = SYNC_FIELDS + [k for k in updates.keys() if k not in SYNC_FIELDS]
    for field in ordered_fields:
        if field.startswith("_"):
            continue
        col = header_map.get(field)
        if not col:
            continue
        value = updates.get(field)
        if value is None:
            if clear_missing:
                ws.cell(row, col).value = None
            continue
        if isinstance(value, str) and not value.strip():
            if clear_missing:
                ws.cell(row, col).value = None
            continue
        ws.cell(row, col).value = value
        written[field] = value
    return written


def main() -> int:
    args = parse_args()
    xlsx_path = Path(args.xlsx).expanduser().resolve()
    cache_path = Path(args.cache).expanduser().resolve()

    run_sync_if_needed(args.no_sync, args.date)
    cache = load_cache(cache_path)
    target_date = pick_target_date(cache, args.date)
    updates = build_updates(
        cache,
        target_date,
        write_summary_to_remark=args.write_summary_to_remark,
        summary_max_len=args.summary_max_len,
        write_inference_reason=not args.no_write_inference_reason,
    )

    if args.query_muscle_group:
        result = {
            "date": target_date,
            "has_training": updates.get("是否训练") == "是",
            "muscle_group": updates.get("训练部位"),
            "workout_summary": updates.get("训练动作"),
            "inference_method": updates.get("_inference_method"),
            "inference_reason": updates.get("_inference_reason"),
            "category_counts": updates.get("_category_counts"),
            "cache_date": updates.get("_cache_date"),
            "cache_match_target": updates.get("_cache_match_target"),
            "allowed_groups": MUSCLE_GROUP_CYCLE,
        }
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return 0

    wb, ws = ensure_workbook(xlsx_path)
    header_row = find_header_row(ws)
    header_map = get_header_map(ws, header_row)
    target_row = find_or_create_date_row(ws, header_row, header_map, target_date)
    written = apply_updates(ws, target_row, header_map, updates, clear_missing=args.clear_missing)
    nutrition_written, diet_status = sync_nutrition_from_row(ws, target_row, header_map)
    written.update(nutrition_written)

    if args.dry_run:
        print(f"[DRY-RUN] xlsx={xlsx_path}")
        print(f"[DRY-RUN] date={target_date}, row={target_row}")
        print(json.dumps(written, ensure_ascii=False, indent=2))
        print(
            f"[DRY-RUN] diet_status: filled_meals={diet_status['filled_meals']}/4, "
            f"summary_present={diet_status['summary_present']}"
        )
        return 0

    wb.save(xlsx_path)
    print(f"✅ 已写入 {xlsx_path}")
    print(f"📅 日期: {target_date} (row {target_row})")
    print("🧾 更新字段:")
    if written:
        for key, value in written.items():
            print(f"  - {key}: {value}")
    else:
        print("  - 无可写入字段（可能缓存数据为空）")
    print(f"🍽 饮食列: {diet_status['filled_meals']}/4")
    print(f"🧠 营养分析列已填写: {'是' if diet_status['summary_present'] else '否'}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print("\n⛔ 已取消")
        raise SystemExit(130)
    except Exception as exc:
        print(f"❌ 同步失败: {exc}")
        raise SystemExit(1)
